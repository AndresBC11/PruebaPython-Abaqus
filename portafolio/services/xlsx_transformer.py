import datetime
from decimal import Decimal

from openpyxl import load_workbook
from portafolio.models import Portafolio, Activos, Activo_en_portafolio, ValorHistoricoPortafolio

def transform_xlsx(ruta_archivo):
    """
    Función para procesar un archivo .xlsx.
    Args: ruta_archivo --> ruta del archivo .xlsx a procesar
    Verifica la existencia de 2 hojas. Hoja 1: Pesos, Hoja 2: Precios. Si no existen, lanza un error.
    """
    print(f"Procesando el archivo: {ruta_archivo}")
    wb = load_workbook(ruta_archivo, data_only=True, read_only=True)
    portafolio_1, _ = Portafolio.objects.get_or_create(nombre="Portafolio 1")
    portafolio_2, _ = Portafolio.objects.get_or_create(nombre="Portafolio 2")
    cargar_activos_y_precios_historicos(wb)
    cargar_pesos_ejercicio(wb, portafolio_1, portafolio_2)
    cantidades_calculadas_1 =calcular_cantidades_historicas(portafolio_1, 1000000000, datetime.date(2022, 2, 15))
    print(f"cantidades_calculadas_1: {cantidades_calculadas_1}")
    cantidades_calculadas_2 =calcular_cantidades_historicas(portafolio_2, 1000000000, datetime.date(2022, 2, 15))
    crear_activos_segun_cantidades_fijas(portafolio_1, wb, cantidades_calculadas_1)
    crear_activos_segun_cantidades_fijas(portafolio_2, wb, cantidades_calculadas_2)
    for fecha in wb['Precios'].iter_rows(min_row=2, values_only=True):
        fecha_actual = fecha[0]
        if fecha_actual:
            calcular_valor_historico_portafolio(portafolio_1, fecha_actual)
            calcular_valor_historico_portafolio(portafolio_2, fecha_actual)
            recalcular_peso_en_activo_en_portafolio(portafolio_1, fecha_actual)
            recalcular_peso_en_activo_en_portafolio(portafolio_2, fecha_actual)

def cargar_activos_y_precios_historicos(wb):
    """
    Función para cargar los activos y sus precios históricos desde un archivo .xlsx.
    Args: ruta_archivo --> ruta del archivo .xlsx a procesar
    """
    print(f"Cargando activos y precios históricos")
    hoja_precios = "Precios"
    
    #Procesamos la hoja de precios
    if hoja_precios in wb.sheetnames:
        hoja = wb[hoja_precios]
        print(f"Procesando hoja de precios")
        nombres_activos = next(hoja.iter_rows(min_row=1, max_row=1, values_only=True))
        activo_actual = 1
        for nombre_activo in nombres_activos[1:]:  # Saltamos la primera columna que es la fecha
            if nombre_activo:
                #Recorremos todas las fechas
                for fecha in hoja.iter_rows(min_row=2, values_only=True):
                    #Para cada fecha revisamos los precios de cada activo y creamos su registro en la base de datos
                    fecha_actual = fecha[0]
                    if fecha_actual:
                        precio = fecha[activo_actual]
                        if precio is not None:
                            activo, _ = Activos.objects.get_or_create(nombre=nombre_activo, valor=precio, fecha=fecha_actual)
            activo_actual += 1
    else:
        raise ValueError(f"La hoja '{hoja_precios}' no existe.")
    

def cargar_pesos_ejercicio(wb, portafolio_1, portafolio_2):
    """
    Función para cargar los pesos de los activos desde un archivo .xlsx.
    Args: wb --> workbook de openpyxl
    """
    print(f"Cargando pesos de los activos")
    hoja_pesos = "weights"
    
    #Procesamos la hoja de pesos
    if hoja_pesos in wb.sheetnames:
        hoja = wb[hoja_pesos]
        print(f"Procesando hoja de pesos")
        #Recorremos todas las fechas
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if fila == (None, None, None, None):
                break  # Salimos del bucle si encontramos una fila vacía
            activo_actual = Activos.objects.filter(nombre=fila[1], fecha=fila[0]).first()
            if activo_actual is None:
                raise ValueError(f"El activo '{fila[1]}' no existe en la base de datos. Asegúrese de que todos los activos estén presentes en la hoja 'Precios' antes de procesar la hoja 'weights'.")
            
            peso_1 = Decimal(str(fila[2]).replace(",", "."))
            peso_2 = Decimal(str(fila[3]).replace(",", "."))
            print(f"Creando o actualizando Activo_en_portafolio para {activo_actual.nombre} en {portafolio_1.nombre} con peso {peso_1}")
            print(f"Creando o actualizando Activo_en_portafolio para {activo_actual.nombre} en {portafolio_2.nombre} con peso {peso_2}")
            
            Activo_en_portafolio.objects.get_or_create(
                portafolio=portafolio_1,
                activo=activo_actual,
                defaults={'peso': peso_1},
            )

            Activo_en_portafolio.objects.get_or_create(
                portafolio=portafolio_2,
                activo=activo_actual,
                defaults={'peso': peso_2}
            )
     
    else:
        raise ValueError(f"La hoja '{hoja_pesos}' no existe en el archivo.")

def calcular_cantidades_historicas(portafolio, valor, fecha):
    """
    Función para calcular las cantidades históricas de los activos en los portafolios.
    """
    print(f"Calculando cantidades históricas de los activos en los portafolios")
    activos_portafolio = Activo_en_portafolio.objects.filter(portafolio=portafolio, fecha=fecha)
    cantidades_calculadas = []
    for activo in activos_portafolio:
        print(f"Calculando cantidad para {activo.activo.nombre} con peso {activo.peso} y valor {valor}")
        if activo.peso > 0:
            cantidad = activo.peso * valor / activo.activo.valor
            cantidad = round(cantidad, 2)
            if cantidad > 0:
                activo.cantidad = cantidad
                activo.valor = activo.activo.valor * cantidad
                activo.save(update_fields=['cantidad', 'valor'])
        cantidades_calculadas.append([activo.activo.nombre, cantidad])
    return cantidades_calculadas

def crear_activos_segun_cantidades_fijas(portafolio, wb, cantidades_calculadas):
    """
    Función para crear activos en portafolio con la misma cantidad historica de una fecha en específico.
    """
    print(f"Asignando cantidades históricas de los activos en los portafolios")
    hoja_precios = "Precios"
    
    #Procesamos la hoja de precios
    if hoja_precios in wb.sheetnames:
        hoja = wb[hoja_precios]
        for fecha in hoja.iter_rows(min_row=2, values_only=True):
            fecha_actual = fecha[0]
            if not fecha_actual:
                continue
            for nombre_activo, cantidad in cantidades_calculadas:
                activo_para_asignar = Activos.objects.filter(nombre=nombre_activo, fecha=fecha_actual).first()
                if activo_para_asignar is None:
                    continue

                Activo_en_portafolio.objects.update_or_create(
                    portafolio=portafolio,
                    activo=activo_para_asignar,
                    defaults={
                        'cantidad': cantidad,
                        'valor': activo_para_asignar.valor * cantidad,
                    }
                )


def calcular_valor_historico_portafolio(portafolio, fecha):
    """
    Función para calcular el valor histórico de un portafolio en una fecha específica.
    """
    print(f"Calculando valor histórico del portafolio '{portafolio.nombre}' en la fecha {fecha}")
    activos_portafolio = Activo_en_portafolio.objects.filter(portafolio=portafolio, fecha=fecha)
    valor_total = 0
    for activo in activos_portafolio:
        valor_total += activo.valor
    ValorHistoricoPortafolio.objects.update_or_create(
        portafolio=portafolio,
        fecha=fecha,
        defaults={
            'valor': valor_total
        }
    )

def recalcular_peso_en_activo_en_portafolio(portafolio, fecha):
    """
    Función para recalcular el peso de cada activo en un portafolio en una fecha específica.
    """
    print(f"Recalculando peso de los activos en el portafolio '{portafolio.nombre}' en la fecha {fecha}")
    activos_portafolio = Activo_en_portafolio.objects.filter(portafolio=portafolio, fecha=fecha)
    valor_historico_portafolio = ValorHistoricoPortafolio.objects.filter(portafolio=portafolio, fecha=fecha).first()
    if valor_historico_portafolio is None:
        raise ValueError(f"No se encontró un valor histórico para el portafolio '{portafolio.nombre}' en la fecha {fecha}.")
    
    for activo in activos_portafolio:
        if valor_historico_portafolio.valor > 0:
            activo.peso = activo.valor / valor_historico_portafolio.valor
        activo.save(update_fields=['peso'])
            